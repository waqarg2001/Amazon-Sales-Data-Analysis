from openpyxl import Workbook, load_workbook
import os
import logging


log_fmt = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
logging.basicConfig(level=logging.INFO, format=log_fmt)


class SingleExcelFile:
    def __init__(self):
        self._logger = logging.getLogger('Sales ETL')

    def multiple_to_single_xlsx_file(self, input_filepath):

        """
        This function takes the directory path of source files and adds them to a new sheet as different
        worksheets. The new excel file is stored in data/interim.

        :param input_filepath: directory path where source files exist
        :return: none
        """
        dest_wb = Workbook()                              # initiating an Excel workbook
        for file in os.listdir(input_filepath):
            if file[-4:] == 'xlsx':
                file = file.replace('.xlsx', '')
                dest_wb.create_sheet(file)                # creating sheet in workbook
                dest_ws = dest_wb[file]
                src_wb = load_workbook("C:\\Users\\HT\\Data Analysis Projects\\Amazon Sales Data\\Amazon Sales Data Analysis\\data\\raw\\"+file+'.xlsx')     # loading workbook
                src_ws = src_wb.active
                for row in src_ws.rows:
                    for cell in row:
                        dest_ws[cell.coordinate] = cell.value
                self._logger.info('Added '+file+'.xlsx '+'to sales_data.xlsx')

        dest_wb.close()
        self._logger.info('Saving sales_data.xlsx...')
        dest_wb.save("C:\\Users\\HT\\Data Analysis Projects\\Amazon Sales Data\\Amazon Sales Data Analysis\\data\\interim\\sales_data.xlsx")


if __name__ == '__main__':
    pass
